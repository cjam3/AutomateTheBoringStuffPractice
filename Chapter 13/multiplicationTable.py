#! python3
# multiplicationTable.py - Creates an N x N multiplication table where N is defined on the command line

import os, sys
import openpyxl
from openpyxl.styles import Font

def main():
    if len(sys.argv) != 2:
        print('Program use --> python3 multiplicationTable.py (N)')
        exit(0)

    createTable(int(sys.argv[1]))


def createTable(tableDim):
    # Create the worksheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Multiplication Table'
    outlineFont = Font(bold = True)
    # Outline the table
    for i in range(2, tableDim + 2):
        sheet.cell(row = 1, column = i).value = i - 1
        sheet.cell(row = 1, column = i).font = outlineFont
        sheet.cell(row = i, column = 1).value = i - 1
        sheet.cell(row = i, column = 1).font = outlineFont

    # Fill in the table
    for i in range(2, tableDim + 2):
        for j in range(2, tableDim + 2):
            sheet.cell(row = i, column = j).value = (i - 1) * (j - 1)

    wb.save('multiplicationTable.xlsx')

if __name__ == '__main__':
    main()