#! Python3
# InvertCell.py - Inverts the cell coordinates in a spreadsheet
# Run in console --> python3 InvertCell.py (name of spreadsheet)

import openpyxl, sys, os
from openpyxl.utils import get_column_letter

def main():
    if len(sys.argv) != 2:
        print('Invalid number of arguments')
        print('Use --> python3 InvertCell.py (name of spreadsheet)')
        exit(0)
    
    InvertCells(sys.argv[1])

def InvertCells(fileName):
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.active
    # Check if the max row or max column is greater
    if sheet.max_column > sheet.max_row:
        greatestDimmension = sheet.max_column
    else:
        greatestDimmension = sheet.max_row

    # Invert cells
    for i in range(1, greatestDimmension + 1):
        for j in range(1, i):
            firstValue = sheet.cell(row = i, column = j).value
            secondValue = sheet.cell(row = j, column = i).value
            sheet.cell(row = i, column = j).value = secondValue
            sheet.cell(row = j, column = i).value = firstValue

    wb.save('InvertCellSpreadsheet_copy.xlsx')


if __name__ == '__main__':
    main()