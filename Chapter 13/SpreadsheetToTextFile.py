#! Python3
# SpreadsheetToTextFile.py - Takes a spreadsheet and converts each column to individual files
# Each row in a column is its own line in the text file
# Use --> python3 SpreadsheetToTextFile.py (name of spreadsheet)

import openpyxl, sys, os
from openpyxl.utils import get_column_letter

def main():
    if len(sys.argv) != 2:
        print('Invalid number of arguments')
        print('Use --> python3 SpreadsheetToTextFile.py (name of spreadsheet)')
        exit(0)

    SpreadsheetToTextFile(sys.argv[1])

def SpreadsheetToTextFile(nameOfFile):
    wb = openpyxl.load_workbook(nameOfFile)
    sheet = wb.active

    # Determine the number of columns and rows 
    numberOfFiles = sheet.max_column
    numberOfLines = sheet.max_row

    for i in range(numberOfFiles):
        # Create file
        fileName = 'Column' + get_column_letter(i + 1) + '.txt'
        textFile = open(fileName, 'w')
        # Write cell value in column to a line in the text file
        for j in range(numberOfLines):
            cell = sheet.cell(row = j + 1, column = i + 1).value
            if cell is None:
                break
            textFile.write(str(cell) + '\n')
        textFile.close()

if __name__ == '__main__':
    main()