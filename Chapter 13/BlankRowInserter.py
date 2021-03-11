#!Python 3
# BlankRowInserter.py - Inserts a specified number of blank rows at a specified row in a spreadsheet
# Run in console python3 BlankRowInserter.py (Row location) (Number of rows) (spreadsheet name)

import openpyxl, sys, os, itertools
from openpyxl.utils import get_column_letter, column_index_from_string

def main():
    if len(sys.argv) != 4:
        print('Invalid number of arguments')
        print('Program use --> python3 BlankRowInserter.py (Row location) (Number of rows) (spreadsheet name)')
        exit(0)

    insertRow(sys.argv[1], sys.argv[2], sys.argv[3])

def insertRow(location, numberOfRows, fileName):
    # Determine the max width of the spreadsheet
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.active
    maxCol = sheet.max_column
    maxRow = sheet.max_row
    # Move the rows down the number of rows starting from the bottom
    for i in range(maxRow, int(location) - 1, -1):
        for j in range(maxCol, 0, -1):
            sheet.cell(row = i + int(numberOfRows), column = j).value = sheet.cell(row = i, column = j).value
    # Create blank rows
    upperLeft = 'A' + location
    bottomRight = get_column_letter(maxCol) + str(int(location) + (int(numberOfRows) - 1))
    sheetSlice = sheet[upperLeft : bottomRight]
    for row in sheetSlice:
        for cell in row:
            cell.value = ''

    wb.save('rowInsertSpreadsheet_copy.xlsx')

if __name__ == '__main__':
    main()